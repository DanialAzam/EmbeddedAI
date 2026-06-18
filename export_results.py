"""Export the benchmark/accuracy results to a formatted Excel workbook.

Reads results/benchmark.csv (and results/accuracy.csv if present) and writes
results/results.xlsx: one sheet per CSV with a styled header, number formats,
auto-filter, frozen header, and a small summary block of MIN/MAX/AVERAGE
formulas. Pure openpyxl - no torch - so it runs on the PC and the Pi.

    python export_results.py
"""
from __future__ import annotations

import csv
from pathlib import Path

# Self-contained paths - deliberately NOT imported from common, which imports cv2
# at module level. The Excel export only needs csv + openpyxl, so it stays light
# and runs under any interpreter (even one without OpenCV/torch).
BASE = Path(__file__).resolve().parent
RESULTS_DIR = BASE / "results"

# number format by case-insensitive header substring
NUMFMT = [("map", "0.0000"), ("mb", "#,##0.00"), ("ms", "0.0"), ("fps", "0.0"),
          ("veh", "0.00"), ("cpu", "0.0"), ("ram", "#,##0")]


def _load(p: Path):
    rows = list(csv.reader(p.read_text(encoding="utf-8").splitlines()))
    return (rows[0], rows[1:]) if rows else ([], [])


def _num(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        return v


def _build_sheet(ws, header, data):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Arial", size=10)
    hdr_fill = PatternFill("solid", fgColor="2563EB")
    edge = Side(style="thin", color="D6DEEA")
    border = Border(left=edge, right=edge, top=edge, bottom=edge)

    for c, h in enumerate(header, 1):
        cell = ws.cell(1, c, h.replace("_", " "))
        cell.font, cell.fill, cell.border = hdr_font, hdr_fill, border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r, row in enumerate(data, 2):
        for c, raw in enumerate(row, 1):
            cell = ws.cell(r, c, _num(raw))
            cell.font, cell.border = body_font, border
            cell.alignment = Alignment(horizontal="left" if c == 1 else "center")
            if isinstance(cell.value, float):
                hl = header[c - 1].lower()
                cell.number_format = next((f for k, f in NUMFMT if k in hl), "0.00")

    for c, h in enumerate(header, 1):
        ws.column_dimensions[get_column_letter(c)].width = max(12, len(h) + 3)
    ws.freeze_panes = "A2"
    if data:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(data) + 1}"


def _summary(ws, header, n):
    """A small block of MIN/MAX/AVERAGE formulas over the data columns."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    def rng(name):
        if name in header:
            L = get_column_letter(header.index(name) + 1)
            return f"{L}2:{L}{n + 1}"
        return None

    items = [("Smallest model size (MB)", "size_mb", "MIN"),
             ("Best max FPS", "max_fps", "MAX"),
             ("Best accuracy (veh/frame)", "veh_per_frame", "MAX"),
             ("Best mAP50-95", "mAP50_95", "MAX"),
             ("Average RAM (MB)", "ram_mb", "AVERAGE")]
    sr = n + 3
    ws.cell(sr, 1, "Summary").font = Font(name="Arial", bold=True, size=11)
    for i, (label, col, fn) in enumerate(items, 1):
        ws.cell(sr + i, 1, label).font = Font(name="Arial", size=10)
        r = rng(col)
        if r:
            cell = ws.cell(sr + i, 2, f"={fn}({r})")
            cell.font = Font(name="Arial", size=10)
            cell.number_format = "0.0000" if col == "mAP50_95" else "#,##0.00"


def main() -> Path:
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    from openpyxl import Workbook

    bench = RESULTS_DIR / "benchmark.csv"
    if not bench.exists():
        raise SystemExit("results/benchmark.csv not found - run 06_benchmark.py first.")

    wb = Workbook()
    header, data = _load(bench)
    ws = wb.active
    ws.title = "Benchmark"
    _build_sheet(ws, header, data)
    _summary(ws, header, len(data))

    acc = RESULTS_DIR / "accuracy.csv"
    if acc.exists():
        h2, d2 = _load(acc)
        if h2:
            _build_sheet(wb.create_sheet("Accuracy (mAP)"), h2, d2)

    out = RESULTS_DIR / "results.xlsx"
    wb.save(out)
    print(f"[xlsx] results -> {out.relative_to(BASE)}  "
          f"({len(data)} models{', + accuracy sheet' if acc.exists() else ''})")
    return out


if __name__ == "__main__":
    main()
